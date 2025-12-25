# D:\New_GAT\core\views\monitoring.py

import json
from collections import defaultdict
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from weasyprint import HTML
from django.template.loader import render_to_string

from .utils_reports import get_report_context
from ..models import SchoolClass

@login_required
def monitoring_view(request):
    """Отображает страницу Мониторинга с новой панелью фильтров."""
    
    # 1. Получаем контекст (фильтры и данные)
    context = get_report_context(request.GET, request.user, mode='monitoring')
    context['title'] = 'Мониторинг'

    # --- 🚀 ОПТИМИЗАЦИЯ: ЗАПРЕТ НА ЗАГРУЗКУ БЕЗ ФИЛЬТРОВ ---
    if not request.GET:
        context['table_rows'] = []   
        context['has_results'] = False 
        context['is_filtered'] = False 
    else:
        context['is_filtered'] = True
        context['has_results'] = bool(context.get('table_rows'))
    # -------------------------------------------------------

    selected_school_ids_str = request.GET.getlist('schools')
    context['selected_class_ids'] = request.GET.getlist('school_classes')
    context['selected_subject_ids'] = request.GET.getlist('subjects')

    context['selected_class_ids_json'] = json.dumps(context['selected_class_ids'])
    context['selected_subject_ids_json'] = json.dumps(context['selected_subject_ids'])

    # --- Группировка классов ---
    grouped_classes = defaultdict(list)
    if selected_school_ids_str:
        try:
            school_ids_int = [int(sid) for sid in selected_school_ids_str]
            classes_qs = SchoolClass.objects.filter(
                school_id__in=school_ids_int
            ).select_related('parent', 'school').order_by('school__name', 'name')

            is_multiple_schools = len(school_ids_int) > 1
            for cls in classes_qs:
                group_name = f"{cls.parent.name} классы" if cls.parent else f"{cls.name} классы (Параллель)"
                if is_multiple_schools:
                    group_name = f"{cls.school.name} - {group_name}"
                grouped_classes[group_name].append(cls)
        except ValueError:
            pass

    final_grouped_classes = {}
    sorted_group_items = sorted(grouped_classes.items(), key=lambda item: (not item[0].endswith("(Параллель)"), item[0]))
    for group_name, classes_in_group in sorted_group_items:
        classes_in_group.sort(key=lambda x: x.name)
        final_grouped_classes[group_name] = classes_in_group

    context['grouped_classes'] = final_grouped_classes

    return render(request, 'dashboard/monitoring.html', context)


@login_required
def export_monitoring_pdf(request):
    """Экспортирует отчет по мониторингу в PDF."""
    context = get_report_context(request.GET, request.user, mode='monitoring')
    context['title'] = 'Отчет по мониторингу'
    
    # Можно добавить сюда ту же логику переводов, если нужно для PDF
    
    html_string = render_to_string('monitoring/monitoring_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="monitoring_report.pdf"'
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(response)
    return response

@login_required
def export_monitoring_excel(request):
    """Экспортирует отчет по мониторингу в Excel с учетом языка."""
    context = get_report_context(request.GET, request.user, mode='monitoring')
    table_headers = context.get('table_headers', [])
    table_rows = context.get('table_rows', [])
    
    # 1. Получаем язык из запроса (по умолчанию 'ru')
    lang = request.GET.get('lang', 'ru')
    
    # 2. Словарь переводов заголовков
    translations = {
        'ru': {
            'no': "№", 'student': "ФИО Студента", 'class': "Класс", 
            'test': "Тест", 'total': "Общий балл", 'from': "из"
        },
        'en': {
            'no': "#", 'student': "Student Name", 'class': "Class", 
            'test': "Test", 'total': "Total Score", 'from': "of"
        },
        'tj': {
            'no': "№", 'student': "Ному насаб", 'class': "Синф", 
            'test': "Тест", 'total': "Холи умумӣ", 'from': "аз"
        }
    }
    # Выбираем текущий словарь, если язык не найден — берем ru
    t = translations.get(lang, translations['ru'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="monitoring_report.xlsx"'
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Monitoring'
    
    # 3. Используем переводы для заголовков
    header1 = [t['no'], t['student'], t['class'], t['test']]
    for header_data in table_headers:
        header1.append(header_data['subject'].abbreviation or header_data['subject'].name)
    header1.append(t['total'])
    
    sheet.append(header1)
    
    header2 = ["", "", "", ""]
    for header_data in table_headers:
        q_count = header_data.get('q_count', 0)
        header2.append(f"({t['from']} {q_count})" if q_count > 0 else "")
    header2.append("")
    
    sheet.append(header2)
    
    # Объединение и стилизация заголовков
    for col in range(1, 5):
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        sheet.cell(row=1, column=col).alignment = Alignment(vertical='center')
    
    total_score_col = len(header1)
    sheet.merge_cells(start_row=1, start_column=total_score_col, end_row=2, end_column=total_score_col)
    sheet.cell(row=1, column=total_score_col).alignment = Alignment(vertical='center')
    
    # 4. Заполняем данными с учетом языка
    for i, row_data in enumerate(table_rows, 1):
        student = row_data['student']
        
        # Выбираем имя в зависимости от языка
        if lang == 'en':
            student_name = student.full_name_en or student.full_name_ru
        elif lang == 'tj':
            student_name = student.full_name_tj or student.full_name_ru
        else:
            student_name = student.full_name_ru

        row = [
            i,
            student_name, # Используем выбранное имя
            str(student.school_class),
            row_data['result_obj'].gat_test.name if row_data.get('result_obj') else "Total"
        ]
        
        for header_data in table_headers:
            score_data = row_data.get('scores_by_subject', {}).get(header_data['subject'].id)
            if score_data and score_data.get('score') != '—':
                cell_value = f"{score_data.get('score', 0)}/{score_data.get('total', 0)}"
            else:
                cell_value = "—"
            row.append(cell_value)
            
        row.append(row_data['total_score'])
        sheet.append(row)
        
    for col_idx, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        sheet.column_dimensions[column].width = (max_length + 2)
        
    workbook.save(response)
    return response