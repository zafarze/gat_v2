# D:\New_GAT\core\views\grading.py (ФИНАЛЬНАЯ ВЕРСИЯ С ПЕРЕВОДОМ ИМЕН)

import json
from collections import defaultdict
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.template.loader import render_to_string
# ✨ 1. ИМПОРТЫ ДЛЯ ЯЗЫКОВ
from django.utils.translation import gettext as _, get_language 
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from .utils_reports import get_report_context
from ..models import SchoolClass

@login_required
def grading_view(request):
    """Отображает страницу 'Таблица оценок'."""
    from django.utils.translation import get_language
    print(f"\n🔥🔥🔥 ПРОВЕРКА ЯЗЫКА: {get_language()} 🔥🔥🔥\n")
    context = get_report_context(request.GET, request.user, mode='grading')
    context['title'] = _('Grading Table')

    selected_school_ids_str = request.GET.getlist('schools')
    context['selected_class_ids'] = request.GET.getlist('school_classes')
    context['selected_class_ids_json'] = json.dumps(context['selected_class_ids'])
    
    context['selected_subject_ids'] = request.GET.getlist('subjects')
    context['selected_subject_ids_json'] = json.dumps(context['selected_subject_ids'])
    
    grouped_classes = defaultdict(list)
    if selected_school_ids_str:
        try:
            school_ids_int = [int(sid) for sid in selected_school_ids_str]
            classes_qs = SchoolClass.objects.filter(
                school_id__in=school_ids_int
            ).select_related('parent', 'school').order_by('school__name', 'name')
            
            is_multiple_schools = len(school_ids_int) > 1
            for cls in classes_qs:
                group_name = f"{cls.parent.name} классы" if cls.parent else f"{cls.name} классы (Параллель)"
                if is_multiple_schools:
                    group_name = f"{cls.school.name} - {group_name}"
                grouped_classes[group_name].append(cls)
        except ValueError:
            pass
    
    final_grouped_classes = {}
    sorted_group_items = sorted(grouped_classes.items(), key=lambda item: (not item[0].endswith("(Параллель)"), item[0]))
    for group_name, classes_in_group in sorted_group_items:
        classes_in_group.sort(key=lambda x: x.name)
        final_grouped_classes[group_name] = classes_in_group
        
    context['grouped_classes'] = final_grouped_classes

    return render(request, 'grading/grading.html', context)

@login_required
def export_grading_pdf(request):
    """Экспортирует отчет по оценкам в PDF."""
    context = get_report_context(request.GET, request.user, mode='grading')
    context['title'] = _('Grading Report')
    html_string = render_to_string('grading/grading_pdf.html', context) 
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="grading_report.pdf"'
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(response)
    return response

@login_required
def export_grading_excel(request):
    """Экспортирует отчет по оценкам в Excel с принудительным выбором языка."""
    # 1. Получаем контекст
    context = get_report_context(request.GET, request.user, mode='grading')
    table_headers = context['table_headers']
    table_rows = context['table_rows']
    
    # ✨ ГЛАВНОЕ ИСПРАВЛЕНИЕ: Читаем язык из ссылки (URL) ✨
    # Если в ссылке есть ?lang=en, то используем его. Если нет — берем системный.
    lang_param = request.GET.get('lang', '').lower()
    
    if lang_param:
        current_lang = lang_param
        print(f"DEBUG: Язык взят из ссылки URL: '{current_lang}'")
    else:
        current_lang = get_language().lower()
        print(f"DEBUG: Язык взят из настроек Django: '{current_lang}'")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="grading_report.xlsx"'
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = str(_('Grading Table'))
    
    # Заголовки
    header1 = [_("№"), _("Student Name"), _("Class"), _("Test")]
    for header_data in table_headers: 
        header1.append(header_data['subject'].abbreviation or header_data['subject'].name)
    header1.append(_("Total Score (grades)"))
    sheet.append(header1)
    
    header2 = ["", "", "", ""]
    for header_data in table_headers: 
        header2.append(_("(10 points)"))
    header2.append("")
    sheet.append(header2)
    
    for col in range(1, 5):
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        sheet.cell(row=1, column=col).alignment = Alignment(vertical='center')
    
    total_score_col = len(header1)
    sheet.merge_cells(start_row=1, start_column=total_score_col, end_row=2, end_column=total_score_col)
    sheet.cell(row=1, column=total_score_col).alignment = Alignment(vertical='center')
    
    for i, row_data in enumerate(table_rows, 1):
        total_grade_score = sum(filter(lambda v: isinstance(v, (int, float)), row_data['grades_by_subject'].values()))
        
        test_name = _("GAT Total") if row_data.get('is_total') else (row_data.get('result_obj').gat_test.name if row_data.get('result_obj') else '')

        # ✨ ЛОГИКА ВЫБОРА ИМЕНИ (С учетом принудительного языка)
        student = row_data['student']
        student_name = student.full_name_ru # По умолчанию
        
        # Проверяем 'en' (включая en-us, en-gb)
        if current_lang.startswith('en'):
            if student.full_name_en and student.full_name_en.strip():
                student_name = student.full_name_en
        # Проверяем 'tj'
        elif current_lang.startswith('tj'):
            if student.full_name_tj and student.full_name_tj.strip():
                student_name = student.full_name_tj
        
        row = [
            i, 
            student_name, 
            str(student.school_class),
            test_name 
        ]
        for header_data in table_headers:
            grade = row_data['grades_by_subject'].get(header_data['subject'].id, "—")
            row.append(grade)
        row.append(total_grade_score)
        sheet.append(row)
        
    for col_idx, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length: 
                    max_length = len(str(cell.value))
            except: 
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
        
    workbook.save(response)
    return response